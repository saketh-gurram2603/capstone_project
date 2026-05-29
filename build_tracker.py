from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Enhancement Tracker"

# ── Palette ──────────────────────────────────────────────────────────────────
HDR_BG   = "1F4E79"
HDR_FG   = "FFFFFF"
HIGH_BG  = "FADADD"
MED_BG   = "FFF2CC"
LOW_BG   = "D9EAD3"
TBD_BG   = "F2F2F2"
OPEN_BG  = "D6E4F0"

CAT_BG = {
    "Authentication":       "EBF5FB",
    "Scalability":          "EBF5FB",
    "Validation":           "EBF5FB",
    "Latency":              "EBF5FB",
    "Session Management":   "EBF5FB",
    "Retry":                "EBF5FB",
    "Feedback & Validation":"FEF9E7",
    "Storage & Persistence":"EAFAF1",
    "Performance":          "EBF5FB",
    "Quality & Accuracy":   "F9EBEA",
    "UX":                   "F5EEF8",
    "Audio Sources":        "FDFEFE",
    "Code & Architecture":  "EBF5FB",
    "Export & History":     "EAFAF1",
    "SSO & Integrations":   "FDF2F8",
}

# ── Data ─────────────────────────────────────────────────────────────────────
rows = [
    # ID  Category                   Feature / Enhancement                      Priority  Description                                                                                                              Status
    (1,  "Authentication",           "User Authentication",                     "TBD",    "Implement secure authentication for the application.",                                                                   "Open"),
    (2,  "Scalability",              "Scalability Testing",                     "High",   "Conduct vertical and horizontal load/scale testing.",                                                                    "Open"),
    (3,  "Validation",               "Google Meet Validation",                  "TBD",    "Validate end-to-end integration with Google Meet.",                                                                      "Open"),
    (4,  "Latency",                  "Latency Measurement & Optimisation",      "TBD",    "Measure and optimise end-to-end latency across the pipeline.",                                                           "Open"),
    (5,  "Storage & Persistence",    "Key Storage",                             "TBD",    "Securely store API keys and credentials.",                                                                               "Open"),
    (6,  "Storage & Persistence",    "Key Encryption",                          "TBD",    "Encrypt keys at rest and in transit.",                                                                                   "Open"),
    (7,  "Session Management",       "Session Termination Policy",              "TBD",    "Define and implement automatic session end when the meeting is over.",                                                   "Open"),
    (8,  "Retry",                    "Retry Logic",                             "TBD",    "Implement retry mechanisms for failed API / network operations.",                                                         "Open"),

    (9,  "Feedback & Validation",    "Post-session Feedback Prompt",            "High",   "Prompt user for rating, correctness score, and free-text notes after each translation session.",                        "Open"),
    (10, "Feedback & Validation",    "MOM Validity Rules",                      "High",   "Validate generated Minutes of Meeting: required sections, action items, decisions, and timestamps must be present.",    "Open"),

    (11, "Storage & Persistence",    "Feedback Storage",                        "High",   "Persist session feedback data for later analysis.",                                                                      "Open"),
    (12, "Storage & Persistence",    "MOM Storage",                             "High",   "Persist generated Minutes of Meeting (MOMs).",                                                                          "Open"),
    (13, "Storage & Persistence",    "Transcript History Storage",              "High",   "Store transcript history; retain 1 month for transcripts, 3 months for MOMs / tasks.",                                  "Open"),

    (14, "Performance",              "Hot-start Time Measurement",              "High",   "Measure and display elapsed time from press-start to first translated text/audio output.",                              "Open"),
    (15, "Performance",              "Startup Optimisation",                    "High",   "Optimise the application startup sequence to reduce cold/warm start latency.",                                          "Open"),
    (16, "Performance",              "Vertical & Horizontal Scale Testing",     "High",   "Run load tests across scaled-up (vertical) and scaled-out (horizontal) configurations.",                               "Open"),

    (17, "Quality & Accuracy",       "Translation Confidence Indicators",       "Low",    "Show confidence / quality indicators where available; target ≥ 80% accuracy threshold.",                               "Open"),
    (18, "Quality & Accuracy",       "Transcription vs Translation Comparison", "High",   "Side-by-side comparison of original transcription and translated text to support quality review.",                     "Open"),

    (19, "UX",                       "Onboarding – Tab Audio Guide",            "High",   "Explain tab audio sharing and required browser actions during first-run onboarding.",                                   "Open"),
    (20, "UX",                       "Onboarding – Flow Diagram",               "High",   "Show a simple step-by-step flow: Select Tab → Share Audio → Translate.",                                               "Open"),
    (21, "UX",                       "Error UI – Failure Reasons",              "High",   "Display clear failure reasons for audio capture, connection, API errors, or missing audio track.",                     "Open"),
    (22, "UX",                       "Error UI – Actionable Suggestions",       "High",   "Provide actionable next-steps when tab audio is unavailable.",                                                          "Open"),
    (23, "UX",                       "Inactivity Notifications",                "High",   "Show notifications or alerts when no audio activity is detected.",                                                      "Open"),
    (24, "UX",                       "Hard Stop – 10 min Inactivity",           "High",   "Automatically terminate the session after 10 consecutive minutes of inactivity.",                                      "Open"),

    (25, "Audio Sources",            "System Mic Live Translation",             "Low",    "Add support for live microphone input as a second audio source.",                                                       "Open"),
    (26, "Audio Sources",            "Audio Source Selection UI",               "Low",    "Let users choose between shared tab audio, system mic, or both.",                                                       "Open"),
    (27, "Audio Sources",            "Mid-session Source Switching",            "Low",    "Allow switching audio source mid-session or launching dual-source mode.",                                               "Open"),

    (28, "Code & Architecture",      "Modularise Client Code",                  "High",   "Refactor monolithic client code into well-defined modules.",                                                            "Open"),
    (29, "Code & Architecture",      "Tests – API Key Flow",                    "High",   "Automated tests covering the API key validation and storage flow.",                                                     "Open"),
    (30, "Code & Architecture",      "Tests – History Save / Load",             "High",   "Automated tests for saving and loading session history.",                                                               "Open"),
    (31, "Code & Architecture",      "Tests – Start / Stop Translation",        "High",   "Automated tests for start and stop translation lifecycle.",                                                             "Open"),
    (32, "Code & Architecture",      "Tests – Error States",                    "High",   "Automated tests covering expected error and failure states.",                                                           "Open"),
    (33, "Code & Architecture",      "Stronger Typing / Linting",               "High",   "Introduce TypeScript strict mode or equivalent linting for safer maintenance.",                                        "Open"),
    (34, "Code & Architecture",      "Constants & Config Isolation",            "High",   "Move repeated strings and config into constants; isolate OpenAI session creation logic.",                              "Open"),

    (35, "Export & History",         "Export – Subtitle Files (.srt/.vtt)",     "High",   "Allow export of session output as .srt or .vtt subtitle files.",                                                       "Open"),
    (36, "Export & History",         "Export – Meeting Notes PDF",              "High",   "Allow export of meeting notes as a formatted PDF.",                                                                     "Open"),
    (37, "Export & History",         "Export – Plain Text",                     "High",   "Allow export of transcript / MOM as plain text.",                                                                      "Open"),
    (38, "Export & History",         "History Filtering & Search",              "High",   "Filter/search history by date, title, language, or feedback score (1 month transcripts, 3 months MOM/tasks).",       "Open"),
    (39, "Export & History",         "Auto-title Preview & Edit",               "High",   "Auto-generate a session title before saving; allow user to edit before confirming.",                                   "Open"),

    (40, "SSO & Integrations",       "SSO + Auto MOM Distribution",             "Low",    "Integrate SSO so the MOM is automatically sent to the meeting participants after the session.",                        "Open"),
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def solid(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def cell_style(ws, row, col, value, bg=None, bold=False, wrap=False,
               font_color="000000", size=10, h_align="left", v_align="top"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", bold=bold, color=font_color, size=size)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if bg:
        c.fill = solid(bg)
    c.border = thin_border()
    return c

# ── Sheet title ───────────────────────────────────────────────────────────────
ws.merge_cells("A1:G1")
t = ws["A1"]
t.value = "Capstone Project – Enhancement & Feature Tracker"
t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
t.fill = solid(HDR_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── Column headers ────────────────────────────────────────────────────────────
headers = ["#", "Category", "Feature / Enhancement", "Priority", "Description", "Status", "Notes"]
for col_idx, h in enumerate(headers, 1):
    cell_style(ws, 2, col_idx, h, bg=HDR_BG, bold=True,
               font_color=HDR_FG, size=10, h_align="center", v_align="center")
ws.row_dimensions[2].height = 22

# ── Data rows ────────────────────────────────────────────────────────────────
for r_offset, (row_id, category, feature, priority, description, status) in enumerate(rows):
    row_num = r_offset + 3
    cat_bg  = CAT_BG.get(category, "FFFFFF")

    pri_bg  = {"High": HIGH_BG, "Low": LOW_BG, "Medium": MED_BG}.get(priority, TBD_BG)
    sta_bg  = OPEN_BG if status == "Open" else "E2EFDA"

    cell_style(ws, row_num, 1, row_id,  bg=cat_bg, h_align="center")
    cell_style(ws, row_num, 2, category, bg=cat_bg, bold=True)
    cell_style(ws, row_num, 3, feature,  bg=cat_bg, wrap=True)
    cell_style(ws, row_num, 4, priority, bg=pri_bg, bold=True, h_align="center")
    cell_style(ws, row_num, 5, description, bg=cat_bg, wrap=True)
    cell_style(ws, row_num, 6, status,   bg=sta_bg, h_align="center")
    cell_style(ws, row_num, 7, "",       bg="FFFFFF")

    ws.row_dimensions[row_num].height = 40

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = [5, 22, 38, 12, 72, 12, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Freeze header rows ────────────────────────────────────────────────────────
ws.freeze_panes = "A3"

# ── Auto-filter on header row ────────────────────────────────────────────────
ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}2"

# ── Legend sheet ─────────────────────────────────────────────────────────────
lg = wb.create_sheet("Legend")
lg.sheet_view.showGridLines = False

legend_data = [
    ("LEGEND",          None,    True,  HDR_BG,  HDR_FG, 12),
    ("Priority Colours","",      True,  "DDEEFF","000000",10),
    ("High",            "",      False, HIGH_BG, "000000",10),
    ("Medium",          "",      False, MED_BG,  "000000",10),
    ("Low",             "",      False, LOW_BG,  "000000",10),
    ("TBD / Unknown",   "",      False, TBD_BG,  "000000",10),
    ("",                "",      False, None,    "000000",10),
    ("Status Colours",  "",      True,  "DDEEFF","000000",10),
    ("Open",            "",      False, OPEN_BG, "000000",10),
    ("Done",            "",      False, "E2EFDA","000000",10),
]

lg.column_dimensions["A"].width = 22
lg.column_dimensions["B"].width = 30

for r, (label, val, bold, bg, fg, sz) in enumerate(legend_data, 1):
    c = lg.cell(row=r, column=1, value=label)
    c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    if bg:
        c.fill = solid(bg)
    c.alignment = Alignment(vertical="center", wrap_text=True)
    lg.row_dimensions[r].height = 18

out_path = r"D:\soft_bank\capstone_project\Enhancement_Tracker.xlsx"
wb.save(out_path)
print("Saved:", out_path)
